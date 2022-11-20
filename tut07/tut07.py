#SIDDHARTH TIWARI 2001CE59
from datetime import datetime
start_time = datetime.now()
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Side
import glob
import os
os.system("cls")







o_s = [1,-1,2,-2,3,-3,4,-4]
onim = {1:"Internal outward interaction", -1:"External outward interaction", 2:"External Ejection", -2:"Internal Ejection", 3:"External inward interaction", -3:"Internal inward interaction", 4:"Internal sweep", -4:"External sweep"}
yellow = "00FFFF00"
yellow_bg = PatternFill(start_color=yellow, end_color= yellow, fill_type='solid')
black = "00000000"
double = Side(border_style="thin", color=black)
black_border = Border(top=double, left=double, right=double, bottom=double)




#Code
def rc(count):
    for item in o_s:
        count[item] = 0

# Method to initialise dictionary with 0 for "o_s" except 'left'
def rce(count, left):
    for item in o_s:
        if(item!=left):
            count[item] = 0

def sf(longest, frequency, outs):
    # Iterating "o_s" and updating sheet
    for i in range(9):
        for j in range(3):
            outs.cell(row = 3+i, column = 45+j).border = black_border

    outs.cell(row=3, column=45).value= "Octant ##"
    outs.cell(row=3, column=46).value= "Longest Subsquence Length"
    outs.cell(row=3, column=47).value= "Count"

    for i, label in enumerate(o_s):
        cR = i+3
        try:
            outs.cell(row=cR+1, column=45).value = label	
            outs.cell(column=46, row=cR+1).value = longest[label]
            outs.cell(column=47, row=cR+1).value = frequency[label]
        except FileNotFoundError:
            print("File not found!!")
            exit()






# Method to set time range for longest subsequence
def longest_subsequence_time(longest, frequency, tr, outs):
    # Naming columns number
    lengthCol = 50
    freqCol = 51
    
    # Initial row, just after the header row
    row = 4

    outs.cell(row=3, column = 49).value = "Octant ###"
    outs.cell(row=3, column = 50).value = "Longest Subsquence Length"
    outs.cell(row=3, column = 51).value = "Count"

    # Iterating all octants 
    for octant in o_s:
        try:
            # Setting octant's longest subsequence and frequency data
            outs.cell(column=49, row=row).value = octant
            outs.cell(column=lengthCol, row=row).value = longest[octant]
            outs.cell(column=freqCol, row=row).value = frequency[octant]
        except FileNotFoundError:
            print("File not found!!")
            exit()

        row+=1

        try:
            # Setting default labels
            outs.cell(column=49, row=row).value = "Time"
            outs.cell(column=lengthCol, row=row).value = "From"
            outs.cell(column=freqCol, row=row).value = "To"
        except FileNotFoundError:
            print("File not found!!")
            exit()

        row+=1

        # Iterating time range values for each octants
        for td in tr[octant]:
            try:
                # Setting time interval value
                outs.cell(row=row, column=lengthCol).value = td[0]
                outs.cell(row=row, column=freqCol).value = td[1]
            except FileNotFoundError:
                print("File not found!!")
                exit()
            row += 1

    for i in range(3, row):
        for j in range(49, 52):
            outs.cell(row=i, column = j).border = black_border



# Method to set frequency count to sheet
			
def fls(outs, t_c):
	# Dictionary to store consecutive sequence count
    count = {}

    # Dictionary to store longest count
    longest = {}

    # Initialing dictionary to 0 for all labels
    rc(count)
    rc(longest)

    # Variable to check last value
    last = -10

    # Iterating complete excel sheet
    for i in range(0, t_c):
        cR = i+3
        try:
            crnt = int(outs.cell(column=11, row=cR).value)

            # Comparing current and last value
            if(crnt==last):
                count[crnt]+=1
                longest[crnt] = max(longest[crnt], count[crnt])
                rce(count, crnt)
            else:
                count[crnt]=1
                longest[crnt] = max(longest[crnt], count[crnt])
                rce(count, crnt)
        except FileNotFoundError:
            print("File not found!!")
            exit()

        # Updating "last" variable
        last = crnt

    # Method to Count longest subsequence frequency
    clsff(longest, outs, t_c)




def clsff(longest, outs, t_c):
    # Dictionary to store consecutive sequence count
    c = {}

    # Dictionary to store frequency count
    fq = {}

    # Dictionary to store time range
    tr = {}

    for l in o_s:
        tr[l] = []

    # Initialing dictionary to 0 for all labels
    rc(c)
    rc(fq)

    # Variable to check last value
    last = -10

    # Iterating complete excel sheet
    for i in range(0, t_c):
        cR = i+3
        try:
            crnt = int(outs.cell(column=11, row=cR).value)
            
            # Comparing current and last value
            if(crnt==last):
                c[crnt]+=1
            else:
                c[crnt]=1        
                rce(c, crnt)

            # Updating freq
            if(c[crnt]==longest[crnt]):
                fq[crnt]+=1

                # Counting starting and ending time of longest subsequence
                end = float(outs.cell(row=cR, column=1).value)
                start = 100*end - longest[crnt]+1
                start/=100

                # Inserting time interval into map
                tr[crnt].append([start, end])

                # Resetting count dictionary
                rc(c)
            else:
                rce(c, crnt)
        except FileNotFoundError:
            print("File not found!!")
            exit()
        except ValueError:
            print("File content is invalid!!")
            exit()

        # Updating 'last' variable
        last = crnt

    # Setting fq table into sheet
    sf(longest, fq, outs)

    # Setting time range for longest subsequence
    longest_subsequence_time(longest, fq, tr, outs)








def tcf(row, T_c, outs):
    # Setting hard coded inputs
    try:
        outs.cell(row=row, column=36).value = "To"
        outs.cell(row=row+1, column=35).value = "Octant #"
        outs.cell(row=row+2, column=34).value = "From"

        for i in range(35, 44):
            for j in range(row+1, row+1+9):
                outs.cell(row=j, column = i).border = black_border

    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    # Setting Labels
    for i, label in enumerate(o_s):
        try:
            outs.cell(row=row+1, column=i+36).value=label
            outs.cell(row=row+i+2, column=35).value=label
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    # Setting data
    for i, l1 in enumerate(o_s):
        maxi = -1

        for j, l2 in enumerate(o_s):
            val = T_c[str(l1)+str(l2)]
            maxi = max(maxi, val)

        for j, l2 in enumerate(o_s):
            try:
                outs.cell(row=row+i+2, column=36+j).value = T_c[str(l1)+str(l2)]
                if T_c[str(l1)+str(l2)] == maxi:
                    maxi = -1
                    outs.cell(row=row+i+2, column=36+j).fill = yellow_bg
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()





def smotc(outs, mod, t_c):
	# Counting partitions w.r.t. mod
    try:
        totP = t_c//mod
    except ZeroDivisionError:
        print("Mod can't have 0 value")
        exit()

    # Checking mod value range
    if(mod<0):
        raise Exception("Mod value should be in range of 1-30000")

    if(t_c%mod!=0):
        totP +=1

    # Initializing row start for data filling
    rS = 16

    # Iterating all partitions 
    for i in range (0,totP):
        # Initializing start and end values
        start = i*mod
        end = min((i+1)*mod-1, t_c-1)

        # Setting start-end values
        try:
            outs.cell(column=35, row=rS-1 + 13*i).value = "Mod Transition Count"
            outs.cell(column=35, row=rS + 13*i).value = str(start) + "-" + str(end)
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        # Initializing empty dictionary
        tCou = {}
        for a in range (1,5):
            for b in range(1,5):
                tCou[str(a)+str(b)]=0
                tCou[str(a)+str(-b)]=0
                tCou[str(-a)+str(b)]=0
                tCou[str(-a)+str(-b)]=0
                
        # Counting transition for range [start, end)
        for a in range(start, end+1):
            try:
                crnt = outs.cell(column=11, row=a+3).value
                next = outs.cell(column=11, row=a+4).value
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

            # Incrementing count for within range value
            if(next!=None):
                tCou[str(crnt) + str(next)]+=1

        # Setting transition counts
        tcf(rS + 13*i, tCou, outs)






def sotc(outs, t_c):
	# Initializing empty dictionary
    c_T = {}
    for i in range (1,5):
        for j in range(1,5):
            c_T[str(i)+str(j)]=0
            c_T[str(i)+str(-j)]=0
            c_T[str(-i)+str(j)]=0
            c_T[str(-i)+str(-j)]=0
        
    # Iterating octants values to fill dictionary
    start = 0

    # try and except block for string to int conversion
    try:
        last = int(outs["K3"].value)
    except ValueError:
        print("Sheet input can't be converted to int")
        exit()
    except TypeError:
        print("Sheet doesn't contain integer octant")
        exit()

    while(start<t_c-1):
        # try and except block for string to int conversion
        try:
            crnt = int(outs.cell(row= start+4, column=11).value)
            c_T[str(last) + str(crnt)]+=1
            last = crnt
        except ValueError:
            print("Sheet input can't be converted to int")
            exit()
        except TypeError:
            print("Sheet doesn't contain integer octant")
            exit()

        start += 1
    
    # Setting transitions counted into sheet
    tcf(2, c_T, outs)






def src(row,countMap, outs):
    # Copying the count list to sort
    sorC = []
    count = []
    for label in o_s:
        count.append(countMap[label])

    for ct in count:
        sorC.append(ct)

    sorC.sort(reverse=True)

    rank = []

    for i, el in enumerate(count):
        for j, ell in enumerate(sorC):
            if(ell==el):
                rank.append(j+1)
                sorC[j] = -1
                break
    r1O = -10

    for j in range(0,8):
        outs.cell(row = row, column=23+j).value = rank[j]
        if(rank[j]==1):
            r1O = o_s[j]
            outs.cell(row = row, column=23+j).fill = yellow_bg    

    outs.cell(row=row , column=31).value = r1O
    outs.cell(row=row , column=32).value = onim[r1O]





def oorf(la_row, outs):
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    row =4
    while outs.cell(row=row, column=29).value is not None:
        oct = int(outs.cell(row=row, column=31).value)
        count[oct]+=1
        row+=1

    for i in range(9):
        for j in range(3):
            row = la_row+2+i
            col = 29+j
            outs.cell(row=row, column = col).border = black_border

    outs.cell(column=29, row=la_row+2).value = "Octant ID"
    outs.cell(column=30, row=la_row+2).value = "Octant Name "
    outs.cell(column=31, row=la_row+2).value = "Count of Rank 1 Mod Values"

    for j, oct in enumerate(o_s):
        outs.cell(column=29, row=la_row+3+j).value = oct
        outs.cell(column=30, row=la_row+3+j).value = onim[oct]
        outs.cell(column=31, row=la_row+3+j).value = count[oct]




def set_mod_count(outs, mod, t_c):
	# Initializing empty dictionary
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    # Variable to store last row
    la_row = -1

    # Iterating loop to set count dictionary
    start = 0
    while(start<t_c):
        try:
            count[int(outs.cell(row=start+3, column=11).value)] +=1
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        start+=1
        try:
            if(start%mod==0):
                # Setting row data
                try:
                    row = 4 + start//mod
                    la_row = row
                    outs.cell(row=row, column=14).value = str(start-mod) + "-" + str(min(t_c, start-1))

                    for i, label in enumerate(o_s):
                        outs.cell(row=row, column=15+i).value = count[label]

                    src(row,count, outs)
                except FileNotFoundError:
                    print("Output file not found!!")
                    exit()
                except ValueError:
                    print("Row or column values must be at least 1 ")
                    exit()

                # Reset count values
                count = {-1:0, 1:0,  -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
        except ZeroDivisionError:
            print("Mod can't have 0 value")
            exit()
    try:
        if(start%mod!=0):
            # Setting row data
            try:
                row = 5 + start//mod
                la_row = row
                outs.cell(row=row, column=14).value = str(start-mod) + "-" + str(min(t_c, start-1))
                for i, label in enumerate(o_s):
                    outs.cell(row=row, column=15+i).value = count[label]
                
                src(row,count, outs)
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

    except ZeroDivisionError:
        print("Mod can't have 0 value")
        exit()

    if(la_row!=-1):
        oorf(la_row, outs)






def sOc(t_c, outs):	
	# Initializing count dictionary
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
    # Incrementing count dictionary data
    try:
        for i in range (3,t_c+3):
            count[int(outs.cell(column=11, row=i).value)] = count[int(outs.cell(column=11, row=i).value)] +1
    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Sheet input can't be converted to int or row/colum should be atleast 1")
        exit()
    except TypeError:
        print("Sheet doesn't contact valid octant value!!")
        exit()

    # Setting data into sheet
    for i, label in enumerate(o_s):
        try:
            outs.cell(row=4, column=i+15).value = count[label]
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    src(4, count, outs)





