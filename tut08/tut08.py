#SIDDHARTH TIWARI 2001CE59
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Border,Side
import numpy as np
import os
import re
from datetime import datetime
os.system('cls')
start_time = datetime.now()

def one_inn(inn1,bat_pl,bow_pl,s): #making function for one innings
    innbat = Workbook()
    innfow = Workbook()
    innbow = Workbook()
    s1 = innbat.active
    s2 = innbow.active
    s3 = innfow.active
    s1.column_dimensions['A'].width = 25

    #scorecard and index
    s1['A1'] = s + ' Innings'
    s1['I1'] = '0-0'
    s1['J1'] = '0 overs'
    s1['A2'] = 'Batter'
    s1['F2'] = 'R'
    s1['G2'] = 'B'
    s1['H2'] = '4s'
    s1['I2'] = '6s'
    s1['J2'] = 'SR'

    s2['A1'] = 'Bowler'
    s2['D1'] = 'O'
    s2['E1'] = 'M'
    s2['F1'] = 'R'
    s2['G1'] = 'W'
    s2['H1'] = 'NB'
    s2['I1'] = 'WD'
    s2['J1'] = 'ECO'

    s3['A1'] = 'Fall of wickets'
    
    #using regex
    over = re.compile(r'(\d\d?\.\d)')
    zero = re.compile(r'no run')
    no_ball = re.compile(r', (no ball),')
    wide = re.compile(r', wide,')
    wide2 = re.compile(r', 2 wides,')
    wide3 = re.compile(r', 3 wides,')
    single = re.compile(r', 1 run,')
    six = re.compile(r', SIX,')
    four = re.compile(r', FOUR,')
    byes = re.compile(r', (byes),')
    lbyes = re.compile(r', (leg byes),')
    double = re.compile(r', 2 runs,')
    triple = re.compile(r', 3 runs,')
    out = re.compile(r', out')
    player = re.compile(r'(\d\d?\.\d) (\w+) (to|\w+ to) (\w+)( \w+)?,')
    caught = re.compile(r', out Caught by (\w+)')
    lbw = re.compile(r', out Lbw!!')
    bowled = re.compile(r', out Bowled!!')
    run_out = re.compile(r'Run Out!! ')
    runs = 0
    wickets = 0
    nb = 0
    nlb = 0
    nw = 0
    nnb = 0
    ppr = 0

    #coding dynamically based on each line
    while True:
        t = inn1.readline()
        if not t:
            break
        crun = 0
        cex = 0
        ov = over.findall(t)
        if float(ov[0]) == int(float(ov[0])) + 0.6:
            ov[0] = str(int(float(ov[0]))+1)
        if float(ov[0])== int(float(ov[0])) + 0.1:
            orun = runs
        nnnb = no_ball.findall(t)
        wbb = wide.findall(t)
        sr = single.findall(t)
        zr = zero.findall(t)
        by = byes.findall(t)
        lby = lbyes.findall(t)
        pl = player.finditer(t)
        w2 = wide2.findall(t)
        w3 = wide3.findall(t)
        for i in pl:
            cbow = i.group(2)
            cbat = i.group(4)
        r = 1
        for col in s1.iter_cols(min_col=1,max_col = 1):
            for cell in col:
                if  cell.value == bat_pl[cbat]:
                    crow = cell.row
                    r = 0
                    break
        if r:
            s1.append([bat_pl[cbat],'not out','','','',0,0,0,0,0])
            crow = s1.max_row

        s = 1
        for col in s2.iter_cols(min_col=1,max_col = 1):
            for cell in col:
                if  cell.value == bow_pl[cbow]:
                    cbrow = cell.row
                    s = 0
                    break
        if s:
            s2.append([bow_pl[cbow],'','',0,0,0,0,0,0,0])
            cbrow = s2.max_row

        #making different cases for different rules in cricket
        if sr:
            crun = 1
        elif zr:
            pass
        else:
            db = double.findall(t)
            if db:
                crun = 2
            else:
                fr = four.findall(t)
                if fr:
                    crun = 4
                    if not (by or lby):
                        s1.cell(row = crow,column = 8).value = s1.cell(row = crow,column = 8).value + 1
                else:
                    sx = six.findall(t)
                    if sx:
                        crun = 6
                        s1.cell(row = crow,column = 9).value = s1.cell(row = crow,column = 9).value + 1
                    else:
                        tp = triple.findall(t)
                        if tp:
                            crun = 3
        if wbb or nnnb or w2 or w3:
            cex = 1
            s2.cell(row = cbrow,column = 6).value = s2.cell(row = cbrow,column = 6).value + 1
            if wbb:
                nw = nw + 1
                s2.cell(row = cbrow,column = 9).value = s2.cell(row = cbrow,column = 9).value + 1 
            elif w2:
                nw = nw + 2
                cex = 2
                s2.cell(row = cbrow,column = 6).value = s2.cell(row = cbrow,column = 6).value + 1
                s2.cell(row = cbrow,column = 9).value = s2.cell(row = cbrow,column = 9).value + 2
            elif w3:
                nw = nw + 3
                cex = 3
                s2.cell(row = cbrow,column = 6).value = s2.cell(row = cbrow,column = 6).value + 2
                s2.cell(row = cbrow,column = 9).value = s2.cell(row = cbrow,column = 9).value + 3
            else:
                nnb = nnb + 1
                s1.cell(row = crow,column = 7).value = s1.cell(row = crow,column = 7).value + 1
                s2.cell(row = cbrow,column = 8).value = s2.cell(row = cbrow,column = 8).value + 1
        else:
            s1.cell(row = crow,column = 7).value = s1.cell(row = crow,column = 7).value + 1
            bo = 10*s2.cell(row = cbrow,column = 4).value -int(s2.cell(row = cbrow,column = 4).value)*4 + 1
            s2.cell(row = cbrow,column = 4).value = int(bo/6)*0.4 + bo*0.1
        runs = runs + crun + cex
        if float(ov[0]) < 6.1:
            ppr = runs
        if float(ov[0])== int(float(ov[0])):
            if orun == runs:
                s2.cell(row = cbrow,column = 5).value = s2.cell(row = cbrow,column = 5).value + 1
        ot = out.findall(t)
        ct = caught.findall(t)
        lw = lbw.findall(t)
        bw = bowled.findall(t)
        ctu = caught.finditer(t)
        ro = run_out.findall(t)
        for i in ctu:
            capl = i.group(1)
        if ot:
            wickets = wickets + 1
            if not ro:
                s2.cell(row = cbrow,column = 7).value = s2.cell(row = cbrow,column = 7).value + 1
            if wickets != 1:
                s3['A2'] = str(s3['A2'].value) + ', ' + str(runs) + '-' + str(wickets) + ' (' + bat_pl[cbat] + ', ' + ov[0] + ')'
            else:
                s3['A2'] = str(runs) + '-' + str(wickets) + '(' + bat_pl[cbat] + ', ' + ov[0] + ')'
            if ct:
                s1.cell(row = crow,column = 2).value = 'c ' + bow_pl[capl] + ' b ' + bow_pl[cbow]
            elif lw:
                s1.cell(row = crow,column = 2).value = 'lbw b ' + bow_pl[cbow]
            elif bw:
                s1.cell(row = crow,column = 2).value = 'b ' + bow_pl[cbow]
            elif ro:
                s1.cell(row = crow,column = 2).value = 'run out (' + bow_pl[cbow] + ')'
            
        #filling in excel
        s1['I1'] = str(runs)+'-'+str(wickets)+'('+ov[0]+' Ov)'
        
        if not (by or lby):
            s1.cell(row = crow,column = 6).value = s1.cell(row = crow,column = 6).value + crun
            s2.cell(row = cbrow,column = 6).value = s2.cell(row = cbrow,column = 6).value + crun
        else:
            if by:
                if fr:
                    nb = nb + 4
                elif db:
                    nb = nb + 2
                elif tp:
                    nb = nb + 3
                elif sr:
                    nb = nb + 1
            else:
                if sr:
                    nlb = nlb + 1
                elif fr:
                    nlb = nlb + 4
                elif db:
                    nlb = nlb + 2
                elif tp:
                    nlb = nlb + 3
            

        s1.cell(row = crow,column = 10).value = float("{:.2f}".format((s1.cell(row = crow,column = 6).value)*100/s1.cell(row = crow,column = 7).value))
        ovf = (10*s2.cell(row = cbrow,column = 4).value - 4*int(s2.cell(row = cbrow,column = 4).value))/6
        if ovf:
            s2.cell(row = cbrow,column = 10).value = float("{:.2f}".format(s2.cell(row = cbrow,column = 6).value/ovf))
        t = inn1.readline()
        if not t:
            break

    #formatting cells
    s1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    s1.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)
    crow = s1.max_row + 1
    s1.merge_cells(start_row=crow, start_column=5, end_row=crow, end_column=7)

    s1.append(['Extras','','','','','','',str(nb + nlb + nw + nnb)+'(b '+str(nb)+', lb '+str(nlb)+', w '+str(nw)+', nb '+str(nnb)+', p 0)'])
    s1.merge_cells(start_row=crow+1, start_column=5, end_row=crow+1, end_column=7)
    s1.append(['Total','','','','','','',str(runs)+'('+str(wickets)+' wkts, '+ov[0]+' Ov)'])
    
    s2.append([])
    s2.append(['Powerplays','Overs','','','','','','','Runs'])
    s2.append(['Mandotary','0.1-6','','','','','','',ppr])

    #combining data
    mc1 = s1.max_column
    mc2 = s2.max_column
    mc3 = s3.max_column
    mr1 = s1.max_row
    mr2 = s2.max_row
    mr3 = s3.max_row


    for i in range(1,mr1-3):
        s1.merge_cells(start_row=i+1, start_column=2, end_row=i+1, end_column=5)
    for i in range(1,mr3+1):
        for j in range(1,9):
            s1.cell(row = mr1+i+1,column=j).value = s3.cell(row = i,column=j).value

    mr11 = s1.max_row
    for i in range(1,mr2+1):
        for j in range(1,11):
            s1.cell(row = mr11+i+3,column=j).value = s2.cell(row = i,column=j).value

    #few more formatting
    s1.merge_cells(start_row=mr1 + 2, start_column=1, end_row=mr1 + 2, end_column=10)
    s1.merge_cells(start_row=mr1 + 3, start_column=1, end_row=mr1 + 5, end_column=10)
    s1.cell(row = mr1+3,column=1).alignment = Alignment(wrap_text=True)
    mr12 = s1.max_row
    s1.merge_cells(start_row=mr12 -1, start_column=2, end_row=mr12-1, end_column=8)
    s1.merge_cells(start_row=mr12, start_column=2, end_row=mr12, end_column=8)
    s1.merge_cells(start_row=mr12 -1, start_column=9, end_row=mr12-1, end_column=10)
    s1.merge_cells(start_row=mr12, start_column=9, end_row=mr12, end_column=10)

    return innbat

